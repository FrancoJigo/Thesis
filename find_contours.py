import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
import os

def image_findContour(im4):
	# load image
	img = cv2.imread(im4)
	gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) # convert to grayscale
	# threshold to get just the signature (INVERTED)
	retval, thresh_gray = cv2.threshold(gray, thresh=100, maxval=255,type=cv2.THRESH_BINARY_INV)

	contours, hierarchy = cv2.findContours(thresh_gray,cv2.RETR_LIST,cv2.CHAIN_APPROX_SIMPLE)

	# # Find object with the biggest bounding box
	# mx = (0,0,0,0)      # biggest bounding box so far
	# mx_area = 0
	# for cont in contours:
	#     x,y,w,h = cv2.boundingRect(cont)
	#     area = w*h
	#     if area > mx_area:
	#         mx = x,y,w,h
	#         mx_area = area
	# x,y,w,h = mx
	# print(len(contours))
	# for h,c in enumerate(contours):
	#     mask = np.zeros(gray.shape,np.uint8)

	#     cv2.drawContours(mask,[c],0,255,-1)
	#     mean = cv2.mean(img,mask = mask)

	if len(contours) != 0:
		# draw in blue the contours that were founded
		cv2.drawContours(img, contours, -1, 255, 3)

		#find the biggest area
		c = max(contours, key = cv2.contourArea)

		x,y,w,h = cv2.boundingRect(c)
		# draw the book contour (in green)
		cv2.rectangle(img,(x,y),(x+w,y+h),(0,255,0),2)
		cv2.imshow("Result", img)
		

	# MAIN PARAMETERS -Feature
	moments = cv2.moments(c)
	#Contour.area - Area bounded by the contour region'''
	area = cv2.contourArea(c)
	# contour perimeter
	perimeter = cv2.arcLength(c,True)
	# centroid
	moments = cv2.moments(c)
	if moments['m00'] != 0.0:
		cx = moments['m10']/moments['m00']
		cy = moments['m01']/moments['m00']
		centroid = (cx,cy)
	else:
		self.centroid = "Region has zero area"
	# bounding box
	bounding_box=cv2.boundingRect(c)
	(bx,by,bw,bh) = bounding_box
	cv2.rectangle(img,(bx,by),(bx+bw,by+bh),(0,255,0),2)
	# aspect ratio
	aspect_ratio = bw/float(bh)
	# equivalent diameter
	equi_diameter = np.sqrt(4*area/np.pi)
	# extent = contour area/boundingrect area
	extent = area/(bw*bh)

	### CONVEX HULL ###
	# convex hull
	convex_hull = cv2.convexHull(c)
	# convex hull area
	convex_area = cv2.contourArea(convex_hull)
	# solidity = contour area / convex hull area
	solidity = area/float(convex_area)

	
	### ELLIPSE  ###
	ellipse = cv2.fitEllipse(c)
	cv2.ellipse(img,ellipse,(50,0,200),2)
	# center, axis_length and orientation of ellipse
	(center,axes,orientation) = ellipse
	# length of MAJOR and minor axis
	majoraxis_length = max(axes)
	minoraxis_length = min(axes)
	# eccentricity = sqrt( 1 - (ma/MA)^2) --- ma= minor axis --- MA= major axis
	eccentricity = np.sqrt(1-(minoraxis_length/majoraxis_length)**2)


	### CONTOUR APPROXIMATION ###
	approx = cv2.approxPolyDP(c,0.02*perimeter,True)

	### EXTREME POINTS ###
	# Finds the leftmost, rightmost, topmost and bottommost points
	leftmost = tuple(c[c[:,:,0].argmin()][0])
	rightmost = tuple(c[c[:,:,0].argmax()][0])
	topmost = tuple(c[c[:,:,1].argmin()][0])
	bottommost = tuple(c[c[:,:,1].argmax()][0])
	extreme = (leftmost,rightmost,topmost,bottommost)

	print('This is the moments',moments)
	print('This is the area',area)
	print(type(area))
	print('This is the perimeter',perimeter)
	print(type(perimeter))
	print('This is the Centroid',centroid)	
	print(type(centroid))
	print('This is the Aspect Ratio',aspect_ratio)
	print(type(aspect_ratio))
	print('This is the equivalent Diameter',equi_diameter)
	print(type(equi_diameter))
	print('This is the Extent',extent)
	print(type(extent))

	print('This is the convex_hull',convex_hull)
	print(type(convex_hull))
	print('This is the convex_area',convex_area)
	print(type(convex_area))
	print('This is the solidity',solidity)
	print(type(solidity))
	print('This is the majoraxis_length',majoraxis_length)
	print(type(majoraxis_length))
	print('This is the minoraxis_length',minoraxis_length)
	print(type(minoraxis_length))
	print('This is the eccentricity',eccentricity)
	print(type(eccentricity))
	print('This is the approx',approx)
	print(type(approx))
	print('This is the leftmost',leftmost)
	print(type(leftmost))
	print('This is the rightmost',rightmost)
	print(type(rightmost))
	print('This is the topmost',topmost)
	print(type(topmost))
	print('This is the bottommost',bottommost)
	print(type(bottommost))
	print('This is the extreme',extreme)
	print(type(extreme))
	
	padding=100
	# Output to files
	roi=img[y-50:y+h+130,x-50:x+w+60]
	# roi=img[y-padding-50:y+h+150,x-padding-50:x+w+150]
	# cv2.imwrite('ROI.jpg', roi)
	cv2.imshow('ROI',roi)
	cv2.rectangle(img,(x,y),(x+w+padding,y+h+padding),(255,0,0),2)
	# cv2.rectangle(img,(x,y),(x+w,y+h),(255,0,0),2)
	# cv2.imwrite('contoured.jpg', img)
	cv2.imshow('Largest contour',img)
	cv2.waitKey(0)
	
	# i = ''
	# wb = load_workbook(filename = "features.xlsx")
	# grab the active worksheet
	# ws = wb.active
	# Data can be assigned directly to cells
	# ws['A'+i] = 9.27
	# ws['B'+i] = area
	# ws['C'+i] = perimeter
	# ws['D'+i] = '('+str(centroid[0])+','+str(centroid[1])+')'
	# ws['E'+i] = aspect_ratio
	# ws['F'+i] = equi_diameter
	# ws['G'+i] = extent
	# # ws['H'+i] = convex_hull
	# ws['H'+i] = convex_area
	# ws['I'+i] = solidity
	# ws['J'+i] = majoraxis_length
	# ws['K'+i] = minoraxis_length
	# ws['L'+i] = eccentricity
	# ws['M'+i] = '('+str(leftmost[0])+','+str(leftmost[1])+')'
	# ws['N'+i] = '('+str(rightmost[0])+','+str(rightmost[1])+')'
	# ws['O'+i] = '('+str(topmost[0])+','+str(topmost[1])+')'
	# ws['P'+i] = '('+str(bottommost[0])+','+str(bottommost[1])+')'	
	# ws['Q'+i] = '('+str(extreme[0])+','+str(extreme[1])+str(extreme[2])+str(extreme[3])+')'
	# ws['N'+i] = approxextreme
	# Save the file
	# wb.save("features.xlsx")
	# os.system("start EXCEL.EXE features.xlsx")

# image_findContour('cleared_border.jpg')